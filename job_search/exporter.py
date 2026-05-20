# =============================================================================
# exporter.py — Export jobs to Excel (.xlsx) with color-coded scores
#               Optionally syncs to Google Sheets via sheets_exporter.py
# =============================================================================

import logging
import os
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

logger = logging.getLogger(__name__)

# ── Google Sheets integration (graceful fallback if unavailable) ───────────────
try:
    from sheets_exporter import export_to_sheets as _sheets_export
    _SHEETS_AVAILABLE = True
except ImportError:
    _SHEETS_AVAILABLE = False
    logger.warning("sheets_exporter not available — Google Sheets sync disabled.")

EXPORTS_DIR = Path(__file__).parent / "exports"
EXPORTS_DIR.mkdir(exist_ok=True)

# ─── Color Palette ────────────────────────────────────────────────────────────
COLORS = {
    "header_bg":    "1A1A2E",   # Dark navy
    "header_fg":    "FFFFFF",
    "subheader":    "16213E",
    "row_even":     "F8F9FF",
    "row_odd":      "FFFFFF",
    "score_A+":     "00C851",   # Green
    "score_A":      "4CAF50",   # Light green
    "score_B":      "FF8800",   # Orange
    "score_C":      "FFD600",   # Yellow
    "score_D":      "FF4444",   # Red
    "link_color":   "0645AD",
    "border":       "D0D0D0",
    "source_naukri":    "E8F5E9",
    "source_linkedin":  "E3F2FD",
    "source_indeed":    "FFF8E1",
    "source_glassdoor": "F3E5F5",
    "source_google":    "FCE4EC",
}

# ─── Column Definitions ───────────────────────────────────────────────────────
COLUMNS = [
    ("Match Score", 12),
    ("Grade",        8),
    ("Job Title",   30),
    ("Company",     22),
    ("Location",    20),
    ("Salary",      18),
    ("Source",      12),
    ("Posted Date", 14),
    ("Apply Link",  50),
]


def export_to_excel(
    jobs: list[dict],
    filename: Optional[str] = None,
    sync_to_sheets: bool = True,
) -> dict:
    """
    Export a list of scored jobs to a beautifully formatted Excel file,
    then optionally sync the same data to Google Sheets.

    Args:
        jobs:           List of scored job dicts from job_searcher.run_job_search()
        filename:       Optional output filename (without extension)
        sync_to_sheets: If True (default), also write to Google Sheets

    Returns:
        Dict with keys:
            'excel_path'  — absolute path to the .xlsx file
            'sheets_url'  — Google Sheets URL (or None if skipped/failed)
    """
    if not jobs:
        raise ValueError("No jobs to export.")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    fname = filename or f"Job_Search_{timestamp}"
    filepath = EXPORTS_DIR / f"{fname}.xlsx"

    wb = openpyxl.Workbook()

    # ─── Main Jobs Sheet (all jobs) ────────────────────────────────────────────
    ws = wb.active
    ws.title = "Job Results"
    _write_jobs_sheet(ws, jobs)

    # ─── Applied / Not Applied Sheets ─────────────────────────────────────────
    applied_jobs     = [j for j in jobs if j.get("applied_status")]
    not_applied_jobs = [j for j in jobs if not j.get("applied_status")]

    ws_applied = wb.create_sheet("Applied Jobs")
    _write_jobs_sheet(ws_applied, applied_jobs)

    ws_not_applied = wb.create_sheet("Not Applied Jobs")
    _write_jobs_sheet(ws_not_applied, not_applied_jobs)

    # ─── Summary Sheet ─────────────────────────────────────────────────────────
    ws_summary = wb.create_sheet("Summary")
    _write_summary_sheet(ws_summary, jobs)

    wb.save(str(filepath))
    logger.info(f"Excel exported: {filepath}")

    # ─── Google Sheets sync ────────────────────────────────────────────────────
    sheets_url = None
    if sync_to_sheets and _SHEETS_AVAILABLE:
        try:
            logger.info("Syncing to Google Sheets...")
            sheets_url = _sheets_export(jobs)
            logger.info(f"Google Sheets updated: {sheets_url}")
        except Exception as exc:
            logger.error(f"Google Sheets sync failed: {exc}")

    return {"excel_path": str(filepath), "sheets_url": sheets_url}


def _write_jobs_sheet(ws, jobs: list[dict]):
    """Write the main job results sheet."""
    # ── Title banner ─────────────────────────────────────────────────────────
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = f"🔍 Job Search Results — {datetime.now().strftime('%d %b %Y, %H:%M')}"
    title_cell.font = Font(name="Calibri", bold=True, size=14, color=COLORS["header_fg"])
    title_cell.fill = PatternFill("solid", fgColor=COLORS["header_bg"])
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:I2")
    subtitle = ws["A2"]
    subtitle.value = f"Total: {len(jobs)} jobs | Showing last 24 hours only | Sorted by Match Score"
    subtitle.font = Font(name="Calibri", size=10, color=COLORS["header_fg"])
    subtitle.fill = PatternFill("solid", fgColor=COLORS["subheader"])
    subtitle.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # ── Column headers ────────────────────────────────────────────────────────
    header_row = 3
    for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.value = col_name
        cell.font = Font(name="Calibri", bold=True, size=10, color=COLORS["header_fg"])
        cell.fill = PatternFill("solid", fgColor="2D3561")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    ws.row_dimensions[header_row].height = 22

    # ── Data rows ────────────────────────────────────────────────────────────
    for row_idx, job in enumerate(jobs, start=header_row + 1):
        is_even = (row_idx % 2 == 0)
        row_bg = COLORS["row_even"] if is_even else COLORS["row_odd"]
        score = job.get("match_score", 0)
        grade = job.get("match_grade", "D")
        source = job.get("source", "").lower()

        # Source-based row tinting
        source_color_map = {
            "naukri": COLORS["source_naukri"],
            "linkedin": COLORS["source_linkedin"],
            "indeed": COLORS["source_indeed"],
            "glassdoor": COLORS["source_glassdoor"],
            "google": COLORS["source_google"],
        }
        row_bg = source_color_map.get(source, row_bg)

        row_data = [
            score,
            grade,
            job.get("title", ""),
            job.get("company", ""),
            job.get("location", ""),
            job.get("salary_string", "N/A") or "N/A",
            source.title(),
            str(job.get("date_posted", "")) or "Today",
            job.get("job_url", ""),
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = _thin_border()
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx in [3, 4, 5]))

            # Score column: color-coded
            if col_idx == 1:
                cell.value = score
                cell.font = Font(name="Calibri", bold=True, size=10,
                                 color=COLORS.get(f"score_{grade}", "000000"))
                cell.fill = PatternFill("solid", fgColor=_score_fill_color(score))
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Grade column
            elif col_idx == 2:
                cell.value = grade
                cell.font = Font(name="Calibri", bold=True, size=10,
                                 color=COLORS.get(f"score_{grade}", "000000"))
                cell.fill = PatternFill("solid", fgColor=_score_fill_color(score))
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Apply Link column — make it a hyperlink
            elif col_idx == 9:
                url = str(value) if value else ""
                if url and url.startswith("http"):
                    cell.value = "Apply →"
                    cell.hyperlink = url
                    cell.font = Font(name="Calibri", size=10, color=COLORS["link_color"],
                                     underline="single")
                else:
                    cell.value = url or "N/A"
                    cell.font = Font(name="Calibri", size=10)
                cell.fill = PatternFill("solid", fgColor=row_bg)

            # Regular columns
            else:
                cell.value = value
                cell.font = Font(name="Calibri", size=10)
                cell.fill = PatternFill("solid", fgColor=row_bg)

        ws.row_dimensions[row_idx].height = 20

    # ── Freeze top 3 rows ────────────────────────────────────────────────────
    ws.freeze_panes = "A4"

    # ── Auto-filter ──────────────────────────────────────────────────────────
    ws.auto_filter.ref = f"A3:I{header_row + len(jobs)}"


def _write_summary_sheet(ws, jobs: list[dict]):
    """Write a summary statistics sheet."""
    ws["A1"] = "📊 Search Summary"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=COLORS["header_bg"])
    ws.merge_cells("A1:D1")
    ws["A1"].alignment = Alignment(horizontal="center")

    # Stats
    grade_counts = Counter(j.get("match_grade", "D") for j in jobs)
    source_counts = Counter(j.get("source", "unknown") for j in jobs)

    stats = [
        ("Total Jobs Found", len(jobs)),
        ("A+ Grade (score ≥80)", grade_counts.get("A+", 0)),
        ("A Grade (score 65–79)", grade_counts.get("A", 0)),
        ("B Grade (score 50–64)", grade_counts.get("B", 0)),
        ("C Grade (score 35–49)", grade_counts.get("C", 0)),
        ("", ""),
        ("Jobs by Source", "Count"),
    ]
    for source, count in source_counts.most_common():
        stats.append((source.title(), count))

    for row_i, (label, value) in enumerate(stats, start=3):
        ws.cell(row=row_i, column=1).value = label
        ws.cell(row=row_i, column=1).font = Font(name="Calibri", bold=bool(label))
        ws.cell(row=row_i, column=2).value = value
        ws.cell(row=row_i, column=2).font = Font(name="Calibri", bold=True)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _thin_border():
    thin = Side(style="thin", color=COLORS["border"])
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _score_fill_color(score: int) -> str:
    """Return a light background fill color based on score range."""
    if score >= 80:
        return "C8F7C5"   # Light green
    elif score >= 65:
        return "D5F5E3"   # Pale green
    elif score >= 50:
        return "FFF3CD"   # Light yellow
    elif score >= 35:
        return "FFE0B2"   # Light orange
    else:
        return "FFCDD2"   # Light red
